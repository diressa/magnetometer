import os
from openpyxl import load_workbook
from statistics import pvariance
from classifier import max_var, max_var2, showplt
from matplotlib import pyplot as plt


def showplt_var(x, y):  # time_tracker.py
    print(file)
    print(max_var(col_y, 100, 50))  # (data, window length, step size)
    markers = max_var2(col_y, 100, 50)  # return just row (sample #, *50)
    markers_x = (sheet_ranges['A' + str(markers)].value, sheet_ranges['A' + str(markers + 100)].value)
    markers_y = (sheet_ranges['C' + str(markers)].value, sheet_ranges['C' + str(markers + 100)].value)
    print(markers, markers_x, markers_y)

    plt.plot(x, y, linewidth=2, color='r')
    plt.plot(markers_x, markers_y, linewidth=2, color='b', marker='x')
    plt.title('Y-Direction of Microteslas')
    plt.xlabel('Time in Seconds')
    plt.ylabel('Microteslas')
    plt.show()


for file in os.listdir('random_xlsx_files'):
    wb = load_workbook(filename='random_xlsx_files/' + file)  # loop through each file
    ws = wb.active

    sheet_ranges = wb['Raw Data']
    data_length = len(sheet_ranges['A'])

    for col_x in ws.iter_cols(min_col=1, min_row=2, max_col=1, max_row=data_length, values_only=True):
        list(col_x)  # print(col_x) # x-axis, time in seconds

    for col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=data_length, values_only=True):
        list(col_y)  # y-axis, magnitude of y-direction microteslas

    showplt_var(col_x, col_y)

'''
Omar's Version, 100 step count limitation:

max_var = 0
num_windows = 200
window_length = 100

for j in range(num_windows):
    list_var = col_y[j*window_length:(j+1)*window_length]
    print(list_var)
    if len(list_var) != 100:
        continue
    current_var = pvariance(list_var)
    max_var = max(current_var, max_var)

print(max_var)
'''

'''
(Instead of only comparing the first sample with the last sample):

see when (time in seconds) the door opens/closes

‘cut’ the start of the file (random number between 1-10 seconds)

& see when the change in action occurs (time in seconds)…

***collect the top 50 samples
***check 0.5-5 second windows (for the sliding window)
***check median/avg of these 100 samples (if 1 sec)
*** windows can overlap (start with 0.5s overlapping windows for example)... to minimize error, 
bc we may not know when action starts exactly 
output when window starts/ends... then compare that number with the actual session where the change starts (ground track)

to see if the program can tell-- not just detecting it at 15 seconds each time

try for at least 5-6 sessions

*or (3/27 meeting) get the average of the first second (100 samples/MHz) and the last second (100 samples/MHz).

'''