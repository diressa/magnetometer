import urllib.request
import time
import webbrowser
import os
from datetime import datetime
import shutil
from classifier import classify, max_var, workbook_run
import xlrd
from openpyxl.workbook import Workbook


IPAddress = '192.168.103.87'                              # Given by PhyPhox app (non-secure http connection)
num_data = 5                                              # Number of data chunks (*100 = # of samples/rows)
save_dat = 'http://' + IPAddress + '/export?format=0'     # Save data format (0: excel, 1: csv inside zip)
clear_dat = 'http://' + IPAddress + '/control?cmd=clear'  # Clear data
start_dat = 'http://' + IPAddress + '/control?cmd=start'  # Start data recording
file_location = '/Users/bezanigatu/Downloads/live_data/'  # Where data is stored


def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)
    os.rename(dst_file_path, dst_file_path + 'x')


def movefile():
    filetime = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
    filename = str('Magnetometer ' + filetime + '.xls')
    time.sleep(5)                                         # wait delay for browser to download
    shutil.move('/Users/bezanigatu/Downloads/' + filename, file_location + filename)
    cvt_xls_to_xlsx(file_location + filename, file_location + filename)  # convert .xls to .xlsx for openpyxl


def restart_data():
    urllib.request.urlopen(clear_dat)                     # Clear then restart data collection
    urllib.request.urlopen(start_dat)


# Data collection begins
urllib.request.urlopen(start_dat)

for v in range(0, num_data):
    webbrowser.open(IPAddress)                            # Uses default browser
    time.sleep(1)                                         # Time in seconds paused * NUM_DATA
    # webbrowser.get("/Applications/Safari.app %s").open(IPAddress)
    # /System/Volumes/Preboot/Cryptexes/App/System/Applications/Safari.app/Contents/MacOS/Safari %s

webbrowser.open_new(save_dat)
time.sleep(0.3)                                           # Re-sync time 0.3 sec to save file
movefile()
restart_data()

for v in range(0, num_data):
    webbrowser.open(IPAddress)
    time.sleep(1)

webbrowser.open_new(save_dat)
time.sleep(0.3)                                           # Re-sync time 0.3 sec to save file
movefile()
urllib.request.urlopen(clear_dat)

# Error check for any rouge excel (.xls) files
stack_file = []

for file in os.listdir(file_location):
    if not file.__contains__('.xlsx'):
        print("Converted file deleted:", file)
        os.remove(file_location + file)
    stack_file.append(file)                             # Order of the directory is consistent, although not 'ordered'


# Classify
stack_classify, stack_nothing, stack_opened, stack_closed = classify('/Users/bezanigatu/Downloads/live_data')

for x in range(0, len(stack_file)-1):
    print(stack_classify[x])
    print(stack_file[x])

for file in os.listdir(file_location):
    if file not in stack_nothing:
        col_x, col_y, sheet_ranges, data_length, wb, ws = workbook_run('/Users/bezanigatu/Downloads/live_data', file)
        max_var(col_y, 100, 50)  # give me times when something is happening

print(stack_classify)


exit()

