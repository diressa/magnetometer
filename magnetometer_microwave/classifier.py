from openpyxl import load_workbook
from statistics import pvariance, mean
from matplotlib import pyplot as plt
import os, random


def showplt(col_x, col_y):      #all_files.py
    plt.plot(col_x, col_y, linewidth=2, color='r')
    plt.title('Y-Direction of Microteslas')
    plt.xlabel('Time in Seconds')
    plt.ylabel('Microteslas')
    plt.show()
    #plt.savefig(file+'.png')


def classify():
    stack_classify = []  # confusion_matrix.py

    stack_opened = []  # display for allfiles.py
    stack_closed = []  # allfiles.py
    stack_nothing = []  # allfiles.py

    for file in os.listdir('random_xlsx_files'):
        wb = load_workbook(filename='random_xlsx_files/' + file)  # loop through each file
        ws = wb.active

        sheet_ranges = wb['Raw Data']
        data_length = len(sheet_ranges['A'])

        for col_x in ws.iter_cols(min_col=1, min_row=2, max_col=1, max_row=data_length, values_only=True):
            list(col_x)  # print(col_x) # x-axis, time in seconds

        for col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=data_length, values_only=True):
            list(col_y)  # y-axis, magnitude of y-direction microteslas μT

        for first_col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=101, values_only=True):
            list(first_col_y)  # collects first second of samples from row 2-101, excluding the header
            tesla_avg1 = mean(first_col_y)  # finds mean average of the first 100 samples
            # print(first_col_y)            # print here to see first 100 samples for all files (20)

        for last_col_y in ws.iter_cols(min_col=3, min_row=(data_length - 100), max_col=3, max_row=data_length,
                                       values_only=True):
            list(last_col_y)  # collects last second of samples, from the last row to 100 before the last row.
            tesla_avg2 = mean(last_col_y)   # finds mean average of the last 100 samples

        # showplt(col_x, col_y)  # allfiles.py

        nothing = 'Nothing'
        opened = 'Close to open'
        closed = 'Open to close'

        if pvariance(col_y) < 0.5:
            classification = nothing
            # print(classification)
            stack_nothing.append(file)
            stack_classify.append(classification)
        elif pvariance(col_y) >= 0.5:
            if tesla_avg1 > tesla_avg2:
                classification = opened
                # print(classification)
                stack_opened.append(file)
                stack_classify.append(classification)
            else:
                classification = closed
                # print(classification)
                stack_closed.append(file)
                stack_classify.append(classification)
    return stack_classify, stack_nothing, stack_opened, stack_closed


stack_classify2, stack_nothing2, stack_opened2, stack_closed2 = classify()


def s_classify():   # confusion_matrix.py
    return stack_classify2


def s_nothing():    # all_files.py
    return stack_nothing2


def s_opened():     # all_files.py
    return stack_opened2


def s_closed():     # all_files.py
    return stack_closed2


# divide data_length/100 for amount of windows
# for x in range(0,data_length,100): #(start, end, iteration) 100 for one-second window, 50 for 0.5 second

def max_var(data, window, step):                     # time_tracker.py
    num_window = (len(data) - window)                # iterate until window size is too small
    variances = []                                   # initialize

    for window_start in range(0, num_window, step):  # increases window's start index by step size 50
        window_end = window_start + window
        window_data = data[window_start:window_end]

        if len(window_data) > 1:                     # does not check (last) window if there's less than 2 data values
            window_variance = pvariance(window_data)
            variances.append(window_variance)

    # print(window_start) to check step sizes
    max_variance = max(variances)
    time_s = 0.5 * variances.index(max_variance)     # only returns the first occurrence of the max value
    max_index = [window_start for window_start, var in enumerate(variances) if var == max_variance]    # list comprehension, checks for duplicate indicies
    print("Time in seconds: ", time_s)               # time_s is (*50 for step size, /100 for sample size), = *0.5

    return max_variance, max_index


def max_var2(data, window, step):                    # time_tracker.py
    num_window = (len(data) - window)                # iterate until window size is too small
    variances = []                                   # initialize

    for window_start in range(0, num_window, step):  # increases window's start index by step size 50
        window_end = window_start + window
        window_data = data[window_start:window_end]

        if len(window_data) > 1:                     # does not check (last) window if there's less than 2 data values
            window_variance = pvariance(window_data)
            variances.append(window_variance)

    # print(window_start) to check step sizes
    max_variance = max(variances)
    sample_start = 50 * variances.index(max_variance)
    # sample_end = 50 * variances.index(max_variance) + 100

    return sample_start  # , sample_end



def workb_ran ():
    sample = random.os.listdir('random_xlsx_files')
    wb = load_workbook(filename='random_xlsx_files/' + sample)  # loop through each file
    ws = wb.active

    sheet_ranges = wb['Raw Data']
    data_length = len(sheet_ranges['A'])

    for col_x in ws.iter_cols(min_col=1, min_row=2, max_col=1, max_row=data_length, values_only=True):
        list(col_x)  # print(col_x) # x-axis, time in seconds

    for col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=data_length, values_only=True):
        list(col_y)  # y-axis, magnitude of y-direction microteslas



