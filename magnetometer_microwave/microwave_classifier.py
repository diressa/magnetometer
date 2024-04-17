from openpyxl import load_workbook
from statistics import pvariance
from matplotlib import pyplot as plt
import os
import random

sample = random.choice(os.listdir('xlsx_files'))
wb = load_workbook(filename='xlsx_files/' + sample)  # loop through each file
ws = wb.active

sheet_ranges = wb['Raw Data']
print(sample)
print(sheet_ranges['A2'].value)

col_x = 0
col_y = 0

for col in ws.iter_cols(min_row=2, max_col=3, max_row=len(sheet_ranges['A']), values_only=True):
    list(col)    # min_row=2 because cannot calculate a mean avg with string in first row

for col_x in ws.iter_cols(min_col=1, min_row=2, max_col=1, max_row=len(sheet_ranges['A']), values_only=True):
    list(col_x)  # x-axis, time in seconds

for col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=len(sheet_ranges['A']), values_only=True):
    list(col_y)  # y-axis, magnitude of y-direction microteslas


print(pvariance(col_y))


plt.plot(col_x, col_y, linewidth=2, color='r')
plt.title('Y-Direction of Microteslas')
plt.xlabel('Time in Seconds')
plt.ylabel('Microteslas')
plt.show()
